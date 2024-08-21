from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.exceptions import ObjectDoesNotExist
from .models import (
    Cpvlicitacion, Codigoscpv, Criterios, Empresas, 
    Licitaciones, Links, Participaciones, Tipocontrato,
    Tipolink, Tipoprocedimiento, Tipotramitacion, Valoraciones
)
from .serializers import (
    CpvlicitacionSerializer, CodigoscpvSerializer, CriteriosSerializer,
    EmpresasSerializer, LicitacionesSerializer,
    LinksSerializer, ParticipacionesSerializer, TipocontratoSerializer,
    TipolinkSerializer, TipoprocedimientoSerializer, TipotramitacionSerializer,
    ValoracionesSerializer
)
import numpy as np
import io
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
import re
from django.db.models import Q
import json
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
import subprocess
import asyncio
from .scripts.main import main
class CpvlicitacionViewSet(viewsets.ModelViewSet):
    queryset = Cpvlicitacion.objects.all()
    serializer_class = CpvlicitacionSerializer

class CodigoscpvViewSet(viewsets.ModelViewSet):
    queryset = Codigoscpv.objects.all()
    serializer_class = CodigoscpvSerializer

class CriteriosViewSet(viewsets.ModelViewSet):
    queryset = Criterios.objects.all()
    serializer_class = CriteriosSerializer

class EmpresasViewSet(viewsets.ModelViewSet):
    queryset = Empresas.objects.all()
    serializer_class = EmpresasSerializer


class LicitacionesViewSet(viewsets.ModelViewSet):
    queryset = Licitaciones.objects.all()
    serializer_class = LicitacionesSerializer

class LinksViewSet(viewsets.ModelViewSet):
    queryset = Links.objects.all()
    serializer_class = LinksSerializer

class ParticipacionesViewSet(viewsets.ModelViewSet):
    queryset = Participaciones.objects.all()
    serializer_class = ParticipacionesSerializer

class TipocontratoViewSet(viewsets.ModelViewSet):
    queryset = Tipocontrato.objects.all()
    serializer_class = TipocontratoSerializer

class TipolinkViewSet(viewsets.ModelViewSet):
    queryset = Tipolink.objects.all()
    serializer_class = TipolinkSerializer

class TipoprocedimientoViewSet(viewsets.ModelViewSet):
    queryset = Tipoprocedimiento.objects.all()
    serializer_class = TipoprocedimientoSerializer

class TipotramitacionViewSet(viewsets.ModelViewSet):
    queryset = Tipotramitacion.objects.all()
    serializer_class = TipotramitacionSerializer

class ValoracionesViewSet(viewsets.ModelViewSet):
    queryset = Valoraciones.objects.all()
    serializer_class = ValoracionesSerializer

class StatisticsView(APIView):
    def get(self, request, licitacion_id):
        try:
            # Get all participaciones related to this licitacion
            participaciones = Participaciones.objects.filter(id_licitacion=licitacion_id)
            
            # Get all valoraciones related to these participaciones
            valoraciones = Valoraciones.objects.filter(id_participacion__in=participaciones)
            # Get unique criterios for which we have valoraciones
            criterios = Criterios.objects.filter(id_criterio__in=valoraciones.values_list('id_criterio', flat=True)).distinct()
            criteriosSumar = Criterios.objects.filter(
                id_criterio__in=valoraciones.values_list('id_criterio', flat=True),
                id_padre__isnull=True
            ).distinct()
            statistics = []
            empresas_total_points = {}
            for participacion in participaciones:
                empresa_id = participacion.id_empresa.id_empresa  # Assuming Participaciones has a foreign key to Empresa
                total_points = 0
                for criterio in criteriosSumar:
                    valoracion = valoraciones.filter(
                        id_participacion=participacion,
                        id_criterio=criterio
                    ).first()
                    if valoracion and valoracion.puntuacion is not None:
                        total_points += valoracion.puntuacion
                
                if empresa_id in empresas_total_points:
                    empresas_total_points[empresa_id] += total_points
                else:
                    empresas_total_points[empresa_id] = total_points

            # Determine the first and second highest total points
            sorted_empresas = sorted(empresas_total_points.items(), key=lambda item: item[1], reverse=True)
            top_empresas = sorted_empresas[:2] if sorted_empresas else []
            if len(top_empresas) == 2:
                top_empresa_1_id = top_empresas[0][0]
                top_empresa_2_id = top_empresas[1][0]
            for criterio in criterios:
                diferencia=""
                if len(top_empresas) == 2:
                    puntuacion_1 = valoraciones.filter(
                            id_participacion__id_empresa=top_empresa_1_id,
                            id_participacion__id_licitacion=licitacion_id,
                            id_criterio=criterio
                        ).first()
                    if puntuacion_1:
                        puntuacion_1=puntuacion_1.puntuacion
                    puntuacion_2 = valoraciones.filter(
                            id_participacion__id_empresa=top_empresa_2_id,
                            id_participacion__id_licitacion=licitacion_id,
                            id_criterio=criterio
                        ).first()
                    if puntuacion_2:
                        puntuacion_2=puntuacion_2.puntuacion
                    if puntuacion_1 and puntuacion_2:
                        diferencia=puntuacion_1-puntuacion_2
                        diferencia=round(diferencia,2)

                # Get puntuaciones for the current criterio
                puntuaciones = valoraciones.filter(id_criterio=criterio).values_list('puntuacion', flat=True)
                puntuaciones = list(puntuaciones)
                median = average = standard_deviation = valor_max=valor_min=0
                if puntuaciones:
                    puntuaciones=[value for value in puntuaciones if value!=None]
                    if puntuaciones!=[]:
                        median = np.median(puntuaciones)
                        average = np.mean(puntuaciones)
                        standard_deviation = np.std(puntuaciones)
                        valor_max = max(puntuaciones)
                        valor_min = min(puntuaciones)
                    
                median = round(median, 2)
                average = round(average, 2)
                standard_deviation = round(standard_deviation, 2)
                statistics.append({
                    'criterio': criterio.nombre,
                    'peso':criterio.valor_max,
                    'inaceptable':criterio.valor_min,
                    'max':valor_max,
                    'min':valor_min,
                    'median': median,
                    'average': average,
                    'standard_deviation': standard_deviation,
                    'id':criterio.id_criterio,
                    'diferencia':diferencia
                })
            

                
                    

            return Response(statistics, status=status.HTTP_200_OK)

        except ObjectDoesNotExist as e:
            return Response({'error': str(e)}, status=status.HTTP_404_NOT_FOUND)
        
        except Exception as e:
            # Log the error for debugging purposes
            import logging
            logging.error(f"An error occurred: {str(e)}")
            return Response({'error': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
def export_licitaciones(request):
    print("Iniciando la exportación de licitaciones.")
    # Obtener datos (ajustar según tu lógica y filtros)
    # Optimizar consultas
    
    
    empresas = Empresas.objects.all()
    filtered_licitaciones = Licitaciones.objects.all().prefetch_related('adjudicatario')
    licitacionesCPV= Cpvlicitacion.objects.all()
    # Crear un libro de trabajo Excel
    wb = Workbook()
    print("Hoja de filtros")
    # Hoja de Filtros
    if 'filtros' in request.GET or 'query' in request.GET:
        names = {
            "lugarEjecucion": "Lugar de Ejecución",
            "importeMax": "Importe Máximo",
            "importeMin": "Importe Mínimo",
            "unidadEncargada": "Unidad Encargada",
            "plazoPresentacionDesde": "Plazo de Presentación (Desde)",
            "plazoPresentacionHasta": "Plazo de Presentación (Hasta)",
            "tipoContrato": "Tipo de Contrato",
            "tipoProcedimiento": "Tipo de Procedimiento",
            "tipoTramitacion": "Tipo de Tramitación",
            "codigoCPV": "Código CPV",
        };
        filtros_ws = wb.create_sheet(title="Filtros")
        filtros_ws.append(["Filtro", "Valor"])
        # Agregar filtros
        if 'query' in request.GET:
            filtros_ws.append(["Búsqueda", request.GET.get('query', "N/A")])
        if 'filtros' in request.GET:
            filtros = request.GET.get('filtros')
            if filtros:
                filtros_dict = json.loads(filtros.strip())
            else:
                filtros_dict = {}
            for filtro in filtros_dict:
                if filtros_dict[filtro] != "":
                    if type(filtros_dict[filtro])==list:
                        for index, cpv in enumerate(filtros_dict[filtro]):
                            filtros_ws.append([f'{names[filtro]} [{index+1}]', f'{cpv["num_cpv"]} - {cpv["descripcion"]}'])
                    else:
                        filtros_ws.append([names[filtro], filtros_dict[filtro]])
        for col in filtros_ws.columns:
            max_length = 30  # Tamaño deseado para todas las columnas
            col_letter = col[0].column_letter
            filtros_ws.column_dimensions[col_letter].width = max_length
    if filtros_dict:
        if 'lugarEjecucion' in filtros_dict and filtros_dict['lugarEjecucion']:
            filtered_licitaciones = filtered_licitaciones.filter(lugar_ejecucion__icontains=filtros_dict['lugarEjecucion'])
        
        if 'importeMax' in filtros_dict and filtros_dict['importeMax']:
            filtered_licitaciones = filtered_licitaciones.filter(importe_sin_impuestos__lte=filtros_dict['importeMax'])
        
        if 'importeMin' in filtros_dict and filtros_dict['importeMin']:
            filtered_licitaciones = filtered_licitaciones.filter(importe_sin_impuestos__gte=filtros_dict['importeMin'])
        
        if 'unidadEncargada' in filtros_dict and filtros_dict['unidadEncargada']:
            filtered_licitaciones = filtered_licitaciones.filter(unidad_encargada__icontains=filtros_dict['unidadEncargada'])
        
        if 'plazoPresentacionDesde' in filtros_dict and filtros_dict['plazoPresentacionDesde']:
            fecha_desde = parse_date(filtros_dict['plazoPresentacionDesde'], "%Y-%m-%d")
            filtered_licitaciones = filtered_licitaciones.filter(plazo_presentacion__gte=fecha_desde)
        
        if 'plazoPresentacionHasta' in filtros_dict and filtros_dict['plazoPresentacionHasta']:
            fecha_hasta = parse_date(filtros_dict['plazoPresentacionHasta'], "%Y-%m-%d")
            filtered_licitaciones = filtered_licitaciones.filter(plazo_presentacion__lte=fecha_hasta)
        
        if 'tipoContrato' in filtros_dict and filtros_dict['tipoContrato']:
            filtered_licitaciones = filtered_licitaciones.filter(tipo_contrato__id_tipo_contrato=filtros_dict['tipoContrato'])
        
        if 'tipoProcedimiento' in filtros_dict and filtros_dict['tipoProcedimiento']:
            filtered_licitaciones = filtered_licitaciones.filter(procedimiento__id_procedimiento=filtros_dict['tipoProcedimiento'])
        
        if 'tipoTramitacion' in filtros_dict and filtros_dict['tipoTramitacion']:
            filtered_licitaciones = filtered_licitaciones.filter(tramitacion__id_tramitacion=filtros_dict['tipoTramitacion'])
        
        if 'codigoCPV' in filtros_dict and filtros_dict['codigoCPV']:
            cpvs = filtros_dict['codigoCPV']
            codigos_cpv = [cpv['num_cpv'] for cpv in cpvs]
            print(codigos_cpv)
            licitaciones_ids = licitacionesCPV.filter(
                id_cpv__num_cpv__in=codigos_cpv
            ).values_list('id_licitacion', flat=True).distinct()
            filtered_licitaciones = filtered_licitaciones.filter(id_licitacion__in=licitaciones_ids)


    
    
    participaciones = Participaciones.objects.filter(
        id_licitacion__in=filtered_licitaciones.values_list('id_licitacion', flat=True)
    ).all()
    valoraciones = Valoraciones.objects.filter(
        id_participacion__in=participaciones.values_list('id_participacion', flat=True)
    ).all()
    print("Hoja de licitaciones")
    # Hoja de Licitaciones
    licitaciones_ws = wb.create_sheet(title="Licitaciones")
    licitaciones_headers = [
        "ID LICITACIÓN", "EXPEDIENTE", "OBJETO", "PLAZO_PRESENTACION", "DURACION",
        "PROCEDIMIENTO", "TRAMITACION", "IMPORTE (SIN IMPUESTOS)", "IMPORTE (CON IMPUESTOS)",
        "VALOR ESTIMADO DEL CONTRATO", "ADJUDICATARIO", "NIF", "¿ES PYME?",
        "OFERTA ADJUDICATARIO", "PUNTUACIÓN ECONÓMICA", "PUNTUACIÓN TÉCNICA", "LUGAR",
        "UNIDAD", "ABONOS A CUENTAS", "TIPO DE CONTRATO", "CLASIFICACION SUBGRUPO",
        "CLASIFICACION GRUPO", "CLASIFICACION CATEGORÍA", "CRITERIOS SOLVENCIA ECONÓMICA",
        "MEDIOS SOLVENCIA ECONÓMICA", "CRITERIOS SOLVENCIA TÉCNICA", "MEDIOS SOLVENCIA TÉCNICA",
        "CRITERIOS PARA DETECTAR VALORES ANORMALES", "CONDICIOENS ESPECIALES",
        "CONSIDERACIÓN COMO INFRACCIÓN GRAVE", "CONTRATACIÓN DEL CONTROL DE CALIDAD",
        "FECHA ADJUDICACIÓN", "FECHA FORMALIZACIÓN", "FECHA ANUNCIO", "FORMA DE PAGO",
        "GARANTÍA DEFINITIVA", "GARANTÍA PROVISIONAL", "GASTOS POR DESISTIMIENTO",
        "MODIFICACIONES PREVISTAS", "PRÓRROGAS PREVISTAS", "REVISIÓN DE PRECIOS PREVISTAS",
        "ORTOS CONCEPTOS PREVISTOS", "INCLUSIÓN DEL CONTROL DE CALIDAD", "MEJORAS COMO CRITERIO",
        "OBLIGACIÓN DE INDICAR SUBCONTRATACIÓN", "OTROS COMPONENTES VALOR ESTIMADO",
        "PENALIDADES POR INCUMPLIMIENTO", "PLAZO DE GARANTÍA", "PLAZO DE RECEPCIÓN",
        "PLAZO MÁXIMO DE LAS PRÓRROGAS", "AMPLIACIÓN DE LA PRESENTACIÓN",
        "POSIBILIDAD DE PRORROGAR EL CONTRATO", "RÉGIMEN DE PENALIDADES", "REVISIÓN DE PRECIOS",
        "SISTEMA DE PRECIOS", "SUBASTA ELECTRÓNICA", "INDICAR SUBCONTRATACIÓN COMO CRITERIO",
        "TAREAS CRÍTICAS",
    ]
    licitaciones_ws.append(licitaciones_headers)

    for licitacion in filtered_licitaciones:
        adjudicatario = licitacion.adjudicatario
        if not adjudicatario:
            continue
        id_empresa = adjudicatario.id_empresa
        id_licitacion = licitacion.id_licitacion
        
        # Buscar la participación del adjudicatario en la licitación
        participacion = participaciones.filter(
            Q(id_licitacion=id_licitacion) & Q(id_empresa=id_empresa)
        ).first()
        valoracionesLicitacion = valoraciones.filter(id_participacion=participacion.id_participacion) if participacion else []

        # Obtener el valor de puntuacion_economica
        punt_economica = next(
            (v for v in valoracionesLicitacion if "OFERTA" in v.id_criterio.nombre.upper() or
            "PRECIO" in v.id_criterio.nombre.upper() or
            "ECONÓMIC" in v.id_criterio.nombre.upper()),
            None
        )

        punt_formulas = next(
            (v for v in valoracionesLicitacion if "FÓRMULA" in v.id_criterio.nombre.upper() or
            "FORMULA" in v.id_criterio.nombre.upper()),
            None
        )

        if punt_economica and punt_formulas and punt_economica.puntuacion and punt_formulas.puntuacion:
            puntuacion_economica = punt_economica.puntuacion + punt_formulas.puntuacion
        else:
            puntuacion_economica = punt_economica.puntuacion if punt_economica else (
                punt_formulas.puntuacion if punt_formulas else ''
            )

        # Obtener el valor de puntuacion_tecnica
        punt_tecnica = next(
            (v for v in valoracionesLicitacion if "JUICIO" in v.id_criterio.nombre.upper() or
            "MEMORIA" in v.id_criterio.nombre.upper()),
            None
        )
        puntuacion_tecnica = punt_tecnica.puntuacion if punt_tecnica else ''
        row = [
            clean_text(licitacion.id_licitacion),  # ID LICITACIÓN
            clean_text(licitacion.num_expediente),  # EXPEDIENTE
            clean_text(licitacion.objeto),  # OBJETO
            clean_text(licitacion.plazo_presentacion),  # PLAZO_PRESENTACION
            clean_text(licitacion.plazo_ejecucion),  # DURACION (Plazo de ejecución)
            clean_text(licitacion.procedimiento.nombre_procedimiento),  # PROCEDIMIENTO
            clean_text(licitacion.tramitacion.nombre_tramitacion) if licitacion.tramitacion else "",  # TRAMITACION
            float(licitacion.importe_sin_impuestos),  # IMPORTE (SIN IMPUESTOS)
            float(licitacion.importe_con_impuestos),  # IMPORTE (CON IMPUESTOS)
            float(licitacion.valor_estimado),  # VALOR ESTIMADO DEL CONTRATO
            clean_text(licitacion.adjudicatario.nombre_empresa),  # ADJUDICATARIO
            clean_text(licitacion.adjudicatario.nif),  # NIF
            "Sí" if licitacion.adjudicatario.pyme else "No",  # ¿ES PYME?
            float(participacion.importe_ofertado_sin_iva) if participacion else '',  # OFERTA ADJUDICATARIO
            puntuacion_economica,  # PUNTUACION ECONOMICA
            puntuacion_tecnica,  # PUNTUACION TECNICA
            clean_text(licitacion.lugar_ejecucion),  # LUGAR
            clean_text(licitacion.unidad_encargada),  # UNIDAD
            clean_text(licitacion.abonos_cuenta),  # ABONOS A CUENTAS
            clean_text(licitacion.tipo_contrato.nombre_tipo_contrato),  # TIPO DE CONTRATO
            clean_text(licitacion.clasificacion_subgrupo),  # CLASIFICACION SUBGRUPO
            clean_text(licitacion.clasificacion_grupo),  # CLASIFICACION GRUPO
            clean_text(licitacion.clasificacion_cat),  # CLASIFICACION CATEGORÍA
            clean_text(licitacion.criterios_economica),  # CRITERIOS SOLVENCIA ECONÓMICA
            clean_text(licitacion.medios_economica),  # MEDIOS SOLVENCIA ECONÓMICA
            clean_text(licitacion.criterios_tecnica),  # CRITERIOS SOLVENCIA TÉCNICA
            clean_text(licitacion.medios_tecnica),  # MEDIOS SOLVENCIA TÉCNICA
            clean_text(licitacion.criterios_valores_anormales),  # CRITERIOS PARA DETECTAR VALORES ANORMALES
            clean_text(licitacion.condiciones_especiales),  # CONDICIOENS ESPECIALES
            clean_text(licitacion.infraccion_grave),  # CONSIDERACIÓN COMO INFRACCIÓN GRAVE
            clean_text(licitacion.contratacion_control),  # CONTRATACIÓN DEL CONTROL DE CALIDAD
            clean_text(licitacion.fecha_adjudicacion),  # FECHA ADJUDICACIÓN
            clean_text(licitacion.fecha_formalizacion),  # FECHA FORMALIZACIÓN
            clean_text(licitacion.fecha_anuncio),  # FECHA ANUNCIO
            clean_text(licitacion.forma_pago),  # FORMA DE PAGO
            clean_text(licitacion.garantia_def),  # GARANTÍA DEFINITIVA
            clean_text(licitacion.garantia_prov),  # GARANTÍA PROVISIONAL
            clean_text(licitacion.gastos_desistimiento),  # GASTOS POR DESISTIMIENTO
            float(licitacion.modificaciones_prev),  # MODIFICACIONES PREVISTAS
            float(licitacion.prorrogas_prev),  # PRÓRROGAS PREVISTAS
            float(licitacion.revision_precios_prev),  # REVISIÓN DE PRECIOS PREVISTAS
            float(licitacion.otros_conceptos_prev),  # OTROS CONCEPTOS PREVISTOS
            clean_text(licitacion.inclusion_control_calidad),  # INCLUSIÓN DEL CONTROL DE CALIDAD
            clean_text(licitacion.mejora_criterio),  # MEJORAS COMO CRITERIO
            clean_text(licitacion.obligacion_subcontratacion),  # OBLIGACIÓN DE INDICAR SUBCONTRATACIÓN
            clean_text(licitacion.otros_componentes),  # OTROS COMPONENTES VALOR ESTIMADO
            clean_text(licitacion.penalidades_incumplimiento),  # PENALIDADES POR INCUMPLIMIENTO
            clean_text(licitacion.plazo_garantia),  # PLAZO DE GARANTÍA
            clean_text(licitacion.plazo_recepcion),  # PLAZO DE RECEPCIÓN
            clean_text(licitacion.plazo_maximo_prorrogas),  # PLAZO MÁXIMO DE LAS PRÓRROGAS
            clean_text(licitacion.ampliacion_presentacion),  # AMPLIACIÓN DE LA PRESENTACIÓN
            clean_text(licitacion.posibilidad_prorroga),  # POSIBILIDAD DE PRORROGAR EL CONTRATO
            clean_text(clean_text(licitacion.regimen_penalidades)),  # RÉGIMEN DE PENALIDADES
            clean_text(licitacion.revision_precios),  # REVISIÓN DE PRECIOS
            clean_text(licitacion.sistema_precios),  # SISTEMA DE PRECIOS
            clean_text(licitacion.subasta_electronica),  # SUBASTA ELECTRÓNICA
            clean_text(licitacion.subcontratacion_criterio),  # INDICAR SUBCONTRATACIÓN COMO CRITERIO
            clean_text(licitacion.tareas_criticas),  # TAREAS CRÍTICAS
        ]
        licitaciones_ws.append(row)
    for col in licitaciones_ws.columns:
        max_length = 30  # Tamaño deseado para todas las columnas
        col_letter = col[0].column_letter
        licitaciones_ws.column_dimensions[col_letter].width = max_length
    print("Hoja de empresas")
    # Hoja de Empresas
    empresas_ws = wb.create_sheet(title="Listado de Empresas")
    empresas_ws.append(["EMPRESAS", "PARTICIPACIONES", "ADJUDICACIONES", "PORCENTAJE DE ÉXITO"])

    for empresa in empresas:
        participaciones_empresa = participaciones.filter(id_empresa=empresa.id_empresa).count()
        adjudicaciones_empresa = filtered_licitaciones.filter(adjudicatario__id_empresa=empresa.id_empresa).count()
        exito = (adjudicaciones_empresa / participaciones_empresa * 100) if participaciones_empresa > 0 else 0
        row = [empresa.nombre_empresa, participaciones_empresa, adjudicaciones_empresa, exito]
        empresas_ws.append(row)
    for col in empresas_ws.columns:
        max_length = 20  # Tamaño deseado para todas las columnas
        col_letter = col[0].column_letter
        empresas_ws.column_dimensions[col_letter].width = max_length
        empresas_ws.column_dimensions["A"].width = 50
    print("Hoja de criterios")
    # Hoja de Criterios
    unique_criterios = set(valoracion.id_criterio.nombre.upper() for valoracion in valoraciones)
    criterios_ws = wb.create_sheet(title="Listado de Criterios")
    criterios_ws.append(["Criterios"])
    numbered_criterios = [(index, criterio) for index, criterio in enumerate(unique_criterios, 1)]
    for index, criterio in numbered_criterios:
        criterios_ws.append([f"{index}-{criterio}"])
    for col in criterios_ws.columns:
        max_length = 200  # Tamaño deseado para todas las columnas
        col_letter = col[0].column_letter
        criterios_ws.column_dimensions[col_letter].width = max_length
    print("Hoja de oferta")
    # Hoja de Oferta Económica
    oferta_ws = wb.create_sheet(title="OFERTA ECONÓMICA")
    oferta_headers = ["EMPRESAS Y PRESUPUESTO"] + [f"LICITACION ({licitacion.id_licitacion})" for licitacion in filtered_licitaciones]
    oferta_ws.append(oferta_headers)
    presupuesto = ["PRESUPUESTO BASE DE LICITACIÓN"]+[float(licitacion.importe_sin_impuestos) for licitacion in filtered_licitaciones]
    oferta_ws.append(presupuesto)
    participaciones_dict = {}
    for participacion in participaciones:
        empresa_id = participacion.id_empresa.id_empresa
        licitacion_id = participacion.id_licitacion.id_licitacion
        importe = participacion.importe_ofertado_sin_iva
        if empresa_id not in participaciones_dict:
            participaciones_dict[empresa_id] = {}
        participaciones_dict[empresa_id][licitacion_id] = importe

    # Construye filas y añade a la hoja
    rows = []
    for empresa in empresas:
        empresa_id = empresa.id_empresa
        row = [empresa.nombre_empresa] + [
            participaciones_dict.get(empresa_id, {}).get(licitacion.id_licitacion, "")
            for licitacion in filtered_licitaciones
        ]
        rows.append(row)

    for row in rows:
        oferta_ws.append(row)
    for col in oferta_ws.columns:
        max_length = 20  # Tamaño deseado para todas las columnas
        col_letter = col[0].column_letter
        oferta_ws.column_dimensions[col_letter].width = max_length
        oferta_ws.column_dimensions["A"].width = 50

    valoraciones_dict = {}
    for valoracion in valoraciones:
        criterio_nombre = valoracion.id_criterio.nombre.upper()
        empresa_id = valoracion.id_participacion.id_empresa.id_empresa
        licitacion_id = valoracion.id_participacion.id_licitacion.id_licitacion
        if criterio_nombre not in valoraciones_dict:
            valoraciones_dict[criterio_nombre]={}
        if empresa_id not in valoraciones_dict[criterio_nombre]:
            valoraciones_dict[criterio_nombre][empresa_id]={}
        valoraciones_dict[criterio_nombre][empresa_id][licitacion_id] = valoracion.puntuacion
    pesos_dict = {}
    for valoracion in valoraciones:
        criterio_nombre = valoracion.id_criterio.nombre.upper()
        criterio_peso = valoracion.id_criterio.valor_max
        licitacion_id = valoracion.id_participacion.id_licitacion.id_licitacion
        if criterio_nombre not in pesos_dict:
            pesos_dict[criterio_nombre]={}
        pesos_dict[criterio_nombre][licitacion_id] = criterio_peso
    # Hoja individual para cada criterio
    for index, criterio in numbered_criterios:
        print(f"{index}-{criterio}")
        rows=[]
        nombre = re.sub(r'[\/\\\*\[\]\:\?]', '_', criterio)
        ws = wb.create_sheet(title=f"{index}-{nombre}")
        headers = ["EMPRESAS Y ESTADÍSTICAS"] + [f"LICITACION ({licitacion.id_licitacion})" for licitacion in filtered_licitaciones]
        ws.append(headers)
        peso = ["PESO"]+[pesos_dict[criterio][licitacion.id_licitacion] if licitacion.id_licitacion in pesos_dict[criterio] else "" for licitacion in filtered_licitaciones] 
        ws.append(peso)
        for empresa in empresas:
            empresa_id = empresa.id_empresa
            row = [empresa.nombre_empresa] + [
                valoraciones_dict.get(criterio, {}).get(empresa_id, {}).get(licitacion.id_licitacion, "")
                for licitacion in filtered_licitaciones
            ]
            
            rows.append(row)
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_length = 20  # Tamaño deseado para todas las columnas
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length
            ws.column_dimensions["A"].width = 50
    print("Guardar")
    # Guardar el workbook en un buffer en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="licitaciones.xlsx"'
    return response

def clean_text(value):
    # Si el valor es None, devuelve una cadena vacía
    if value is None:
        return ''
    
    # Convertir el valor a string
    value = str(value)
    
    # Eliminar caracteres de control que no son permitidos
    value = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
    
    # Limitar la longitud del texto a un máximo, por ejemplo, 32,767 caracteres (límite de Excel)
    max_length = 32767
    if len(value) > max_length:
        value = value[:max_length]
    
    return value

def parse_date(date_str, fmt):
    """ Convierte una fecha en formato de cadena a un objeto datetime.date """
    try:
        return datetime.strptime(date_str, fmt).date()
    except ValueError:
        return None
    
@csrf_exempt
def run_script_extraccion(request):
    if request.method == 'POST':
        try:
            results = asyncio.run(main()) 
            return JsonResponse({'status': 'success', 'data': results})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)

class CPVListView(APIView):
    def get(self, request, licitacion_id):
        try:
            # Buscar todas las entradas de CPVLicitacion para el ID de licitación dado
            cpv_licitaciones = Cpvlicitacion.objects.filter(id_licitacion=licitacion_id)
            
            # Extraer los códigos CPV
            cpv_codigos = [(cpv_licitacion.id_cpv.num_cpv, cpv_licitacion.id_cpv.descripcion) for cpv_licitacion in cpv_licitaciones]

            return JsonResponse({'status': 'success', 'cpv_codigos': cpv_codigos})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)